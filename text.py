import openpyxl


def exporting(address, street):
    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    count = 1
    c1 = my_sheet.cell(row = int(count), column = 1)
    c1.value = address
    c2 = my_sheet.cell(row = int(count), column = 2)
    c2.value = street
    count += 1
    my_wb.save (r"C:\Users\big shot\Documents\GitHub\parcing\2.xlsx")

address, street = "lola", "kok"

exporting(address, street)
