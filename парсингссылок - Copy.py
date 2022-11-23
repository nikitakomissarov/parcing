
import urllib.request, urllib.parse, urllib.error
from bs4 import BeautifulSoup
import ssl
import re
import openpyxl


ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

url = input('Enter URL ')


html = urllib.request.urlopen(url, context=ctx).read()
soup = BeautifulSoup(html, 'html.parser')
tags = soup('a')

def finding(neededmeta, address):
    coordinates = re.findall(r'[0-9][0-9][.][0-9]+', neededmeta)
    latitude = coordinates[0]
    longitude = coordinates[1]
    print(latitude, longitude)
    parcing2(address, timer, latitude, longitude)

def exporting(newaddress, street, timer, latitude, longitude, my_wb):
    my_sheet = my_wb.active
    c1 = my_sheet.cell(row = int(timer), column = 1)
    c1.value = newaddress
    c2 = my_sheet.cell(row = int(timer), column = 2)
    c2.value = street
    c3 = my_sheet.cell(row = int(timer), column = 3)
    c3.value = longitude
    c4 = my_sheet.cell(row = int(timer), column = 4)
    c4.value = latitude
    print('TIMER', timer)

def parcing2(address, timer, latitude, longitude):
    b = (str(address)).upper().split(' ')
    newaddress, street = b[0], b[-1]
    exporting(newaddress, street, timer, latitude, longitude, my_wb)
    print(newaddress, street, '\n')
    
my_wb = openpyxl.Workbook()
    
timer = 1
count = 0
for link in tags:
    count += 1
    add = link.string
    link = link.get('href')
    nextlink = "https://mapdata.ru" + link
    print(nextlink)
    if not nextlink.startswith("https://mapdata.ru/kaliningradskaya-oblast/"): continue
    print('COUNT', count)
    nexthtml = urllib.request.urlopen(nextlink, context=ctx).read()
    newsoup = BeautifulSoup(nexthtml, 'html.parser')
    newtag = newsoup('meta')
    if len(newtag) < 5: continue
    neededmeta = newtag[6]
    address = add
    finding(str(neededmeta), address)
    timer += 1 

my_wb.save (r"C:\Users\Никита\Desktop\Book7.xlsx") 
