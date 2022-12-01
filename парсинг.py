import urllib.request, urllib.parse, urllib.error
from bs4 import BeautifulSoup
import ssl
import openpyxl


# Ignore SSL certificate errors
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

url = input('Enter URL ')
b = list()

html = urllib.request.urlopen(url, context=ctx).read()
soup = BeautifulSoup(html, 'html.parser')
tags = soup('a')



v = open('location1.txt', 'w')
f = open('location2.txt', 'w')

my_wb = openpyxl.Workbook()

def exporting(address, street, count, my_wb):
    my_sheet = my_wb.active
    c1 = my_sheet.cell(row = int(count), column = 1)
    c1.value = address
    c2 = my_sheet.cell(row = int(count), column = 2)
    c2.value = street
    print('COUNT', count)


def parcing(tags):
    count = 1
    for tag in tags:
        count += 1
        a = tag.string
        a = (str(a)).upper().split(',')
        address, street = a[0], a[-1]
        #a = a.split(',')
        exporting(address, street, count, my_wb)
        v.write(address + '\n')
        f.write(street + '\n')
        print(address, street)
        
parcing(tags)

my_wb.save (r"C:\Users\big shot\Documents\GitHub\parcing\2.xlsx")

v.close()
f.close()
