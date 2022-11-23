import urllib.request, urllib.parse, urllib.error
from bs4 import BeautifulSoup
import ssl
import openpyxl


# Ignore SSL certificate errors
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

url = input('Enter URL ')
b = list()

html = urllib.request.urlopen(url, context=ctx).read()
soup = BeautifulSoup(html, 'html.parser')
tags = soup('a')


v = open('location1.txt', 'w')
f = open('location2.txt', 'w')


    


def parcing(tags):
    for tag in tags:
        a = tag.string
        a = (str(a)).upper().split(',')
        address, street = a[0], a[-1]
        #a = a.split(',')
        my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active
        count = 1
        c1 = my_sheet.cell(row = count), column = 1)
        c1.value = address
        c2 = my_sheet.cell(row = count), column = 2)
        c2.value = street
        count += 1
        my_wb.save (r"C:\Users\Никита\Desktop\Book5.xlsx")
        v.write(address + '\n')
        f.write(street + '\n')
        print(address, street)

parcing(tags)    

v.close()
f.close()
