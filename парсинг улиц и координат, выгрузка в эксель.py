import urllib.request, urllib.parse, urllib.error
from bs4 import
import ssl
import openpyxl


# Ignore SSL certificate errors
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

url = input('Enter URL ')
b = list()

html = urllib.request.urlopen(url, context=ctx).read()
soup = BeautifulSoup(html, 'html.parser')
tags = soup('a')

v = open('location1.txt', 'w')
f = open('location2.txt', 'w')

my_wb = openpyxl.Workbook()

def exporting(address, street, count, my_wb):
    my_sheet = my_wb.active
    c1 = my_sheet.cell(row = int(count), column = 1)
    c1.value = address
    c2 = my_sheet.cell(row = int(count), column = 2)
    c2.value = street
    print('COUNT', count)

def parcing2(retrievedstr, count):
    b = (str(retrievedstr)).upper().split(' ')
    address, street = b[0], b[-1]
    exporting(address, street, count, my_wb)
    v.write(address + '\n')
    f.write(street + '\n')
    print(address, '\n', street)



def parcing1(retrievedstr2, count):
    address, street = retrievedstr2[0], retrievedstr2[-1]
    exporting(address, street, count, my_wb)
    v.write(address + '\n')
    f.write(street + '\n')
    print(address, '\n', street) 
 

def parcing(tags):
    count = 0
    for tag in tags:
        retrievedstr = tag.string
        if retrievedstr == None: continue
        retrievedstr2 = (str(retrievedstr)).upper().split(',')
        count += 1
        if retrievedstr2[-1] == retrievedstr2[0]:
            parcing2(retrievedstr, count)    
        else:
            parcing1(retrievedstr2, count)

    

           
parcing(tags)
my_wb.save (r"C:\Users\Никита\Desktop\Book7.xlsx") 

v.close()
f.close()
