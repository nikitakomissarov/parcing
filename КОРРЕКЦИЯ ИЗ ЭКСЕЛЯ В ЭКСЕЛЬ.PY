import openpyxl

command = input()
namecount = 0
while command != 'stop' or 'стоп':
    name = input("Введите путь файла")
    my_path = ( name)
# C:\Users\Никита\Desktop\КАЛИНИНГРАД1.xlsx
# C:\Users\Никита\Desktop\БАГРАТИОНОВСК.xlsx
# C:\Users\Никита\Desktop\ГУСЕВ1.xlsx
# C:\Users\Никита\Desktop\БАЛТИЙСК1.xlsx
# C:\Users\Никита\Desktop\ЗЕЛЕНОГРАДСК1.xlsx
# C:\Users\Никита\Desktop\ГУРЬЕВСК1.xlsx
# C:\Users\Никита\Desktop\МАМОНОВО1.xlsx
# C:\Users\Никита\Desktop\СОВЕТСК1.xlsx
# C:\Users\Никита\Desktop\СВЕТЛЫЙ.xlsx
# C:\Users\Никита\Desktop\ЯНТАРНЫЙ.xlsx
# C:\Users\Никита\Desktop\КРАСНОЗНАМЕНСК.xlsx
# C:\Users\Никита\Desktop\НЕМАН.xlsx
# C:\Users\Никита\Desktop\СВЕТЛОГОРСК1.xlsx
# C:\Users\Никита\Desktop\ОЗЕРСК.xlsx
# C:\Users\Никита\Desktop\ЧЕРНЯХОВСК1.xlsx
# C:\Users\Никита\Desktop\ЛАДУШКИН.xlsx
# C:\Users\Никита\Desktop\НЕСТЕРОВ.xlsx
# C:\Users\Никита\Desktop\СЛАВСК.xlsx
# C:\Users\Никита\Desktop\ПОЛЕССК.xlsx
# C:\Users\Никита\Desktop\ПРАВДИНСК.xlsx
# C:\Users\Никита\Desktop\ПИОНЕРСКИЙ.xlsx

    my_wb = openpyxl.load_workbook(my_path)
    my_sheet = my_wb.active
    c1  = my_sheet.cell(row = 1,column = 1)
    c2  = my_sheet.cell(row = 1,column = 2)

    count = 3

    prestreet = ['УЛИЦА', 'ПЕРЕУЛОК', 'ПЛОЩАДЬ', 'ПРОЕЗД', 'БУЛЬВАР', 'ПРОСПЕКТ', 'ТУПИК', 'АЛЛЕЯ', 'НАБЕРЕЖНАЯ', 'СПУСК']

    
    while c1.value != None:
        count += 1
        c1 = my_sheet.cell(row = count, column = 1)
        c2  = my_sheet.cell(row = count,column = 2)
        if c1.value in prestreet: continue
        bufer = c2.value
        c2.value = c1.value
        c1.value = bufer
        print(c1.value, c2.value)
        print('COUNT', count)
        if c1.value == None:
            print("KONЕЦ")

    namecount += 1
    nameend = "NEWНЕСТЕРОВ_" + str(namecount)
    my_wb.save(r'C:\Users\Никита\Desktop\bot\ '+ nameend)


